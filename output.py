from network import Network
from typing import Sequence, TypeAlias, Mapping, NamedTuple, Final, TypedDict
from readxl import read 
from openpyxl import Workbook, worksheet
import openpyxl
from dataclasses import dataclass, field
import sys 
from warnings_ import NotAllData

parameter: TypeAlias = str 
sensitivity: TypeAlias = float
id: TypeAlias = int 
NetworkSequence: TypeAlias = Sequence[Network, ]

LANG: Final = sorted('QWERTYUIOPASDFGHJKLMNBVCXZ')
NORMAL_COLOR: Final = '000000'
SELECTED_COLOR: Final = 'FFAA00'


class Cell(NamedTuple): 
    data: str | int | float 
    font: openpyxl.styles.Font = openpyxl.styles.Font(color=NORMAL_COLOR, 
                                                      bold=False, 
                                                      italic=False, 
                                                      size=12)

class SortedSensitivity(NamedTuple): 
    param: Sequence[str, ]
    sensitivity: Sequence[int | float, ]
    pasition: Sequence[int, ]

@dataclass
class Row: 
    parameter: str=''
    sensitivity: list[Cell, ] = field(init=False, repr=False, default_factory=list)
    count: int = 0

def get_sorted_Sensativity(networks: NetworkSequence) -> dict[id, SortedSensitivity]: 
    out = dict()
    for net in networks: 
        sensitivity = net.sensitivity
        sensitivity = sorted(sensitivity, key=lambda x: x[1], reverse=True)

        param = [x[0] for x in sensitivity]
        sensitivity = [x[1] for x in sensitivity]
        ids = list(range(1, len(sensitivity)+1))

        out[net.id] = SortedSensitivity(param=param, sensitivity=sensitivity, pasition=ids)
    return out
    


def __get_the_worst_param(net:Network, count:int=6) -> Mapping[parameter, sensitivity]: 
    """Фунция, возрващающая count намиенее значимых параметров в конкретной нейронной сети"""
    parameters = net.sensitivity
    parameters = sorted(parameters, key=lambda x: x[1])
    return dict(parameters[:count])

def get_dict_the_worst_param(networks: NetworkSequence) -> dict[id, tuple[str,]]: 
    """Функция, возвращающая список нименее чувствительных параметров для каждый сети в списке"""
    network_dict = dict() 
    for net in networks: 
        the_worst = __get_the_worst_param(net)                                      # на выход получаем список из картежей (параметр, чувствительность)
        network_dict[net.id] = the_worst.keys()                                     # переводим в формат славаря 

    return network_dict
    
def get_sheet(networks: NetworkSequence) -> Sequence[Sequence[Cell,], ]: 
    global NotAllData
    sheet:list[Row, ] = []                                                          
    parameters = list(map(lambda x: x[0], networks[1].sensitivity))                              
    ids = sorted(map(lambda x: x.id, networks), key=int)   
    sortedSensitivity  = get_sorted_Sensativity(networks)            

    header = list(map(Cell, ["Параметр", 
                             *tuple(map(lambda x: "Сеть " + str(x), ids)), 
                             "Сумма позиций"]))

    for param in parameters:                                        
        try:
            row = __get_row(networks, param, sortedSensitivity)
        except IndexError: 
            print("Данные не полны", file=sys.stderr)
            NotAllData = "Пропущены данные в таблице чувтвительности"
        else: 
            sheet += [[row.parameter, *row.sensitivity, row.count]]

    return [header] + sorted(sheet, key=lambda x: x[-1].data, reverse=True)         

def __get_row(networks:NetworkSequence, param:str, sortedSensitivity: dict[id, SortedSensitivity]) -> Row: 
    """Функция, возвращающая строку таблицы"""
    row = Row(parameter=Cell(param))
    for net in networks: 
        id = net.id
        for parameter, sens, position in zip(*sortedSensitivity[id]): 
            if parameter == param: 
                row.count += position
                row.sensitivity += [Cell(sens)]
                break 

    row.count = Cell(row.count)
    return row 

def get_norm_pos(a:int) -> str: 
    """Функция перевода числа в вид столбца Excel"""
    out = ''
    while a: 
        out += LANG[((a-1)%len(LANG))]
        a = (a-1)//len(LANG)
    return out[::-1]

def write_sheet(ws: worksheet, sheet: Sequence[Sequence[Cell, ], ]) -> None: 
    column_index = 1
    row_index = 1
    for row in sheet: 
        for cell in row: 
            column = get_norm_pos(column_index)
            ws[column+str(row_index)].font = cell.font
            ws[column+str(row_index)] = cell.data
            column_index += 1 
        else:
            column_index = 1
            row_index += 1 

def create_work_book(networks: NetworkSequence, name:str='Анализ') -> Workbook: 
    wb = Workbook()
    ws = wb.active
    ws.title = name
    sheet = get_sheet(networks)
    write_sheet(ws, sheet)
    return wb 

if __name__ == "__main__": 
    nets = read('outt.xlsx')
    wb = create_work_book(nets)
    wb.save("out.xlsx")
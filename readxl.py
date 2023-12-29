from openpyxl import load_workbook, Workbook, worksheet
from dataclasses import dataclass, field
from typing import Sequence, TypeAlias, Mapping
from network import Network
from exceptions import IncorrectFile
from warnings_ import *
from itertools import zip_longest

id: TypeAlias = int 


def read(filename:str) ->  Sequence[Network, ]:  
    """Функция четения данных их файла Excel"""
    workbook = load_workbook(filename)
    networks = dict()   
    if len(workbook.sheetnames) < 2: 
        raise IncorrectFile("В файле должно быть 2 листа. В первом чувтвительность, а во втором кофигурация сетей")

    networks = get_networks_sesensitivity(workbook[workbook.sheetnames[0]], networks)
    get_networks_config(workbook[workbook.sheetnames[1]], networks)

    return list(networks.values())
    

def get_networks_config(ws: worksheet, networks: Mapping[id, Network]) -> Mapping[id, Network]:
    """Функция сбора информации о сетях из файла xlsx"""
    global ConfigNotAllData
    rows = tuple(ws.values)[1:]
    for row in map(tuple, rows):
        id, arhitecture, training_performance, validate_performance,  test_performance, training_erro, validate_error, \
             test_error, learning_alghoritm, error_function, activate1, activate2 = row
        
        if id is None: 
            continue

        if id in networks:
            network = networks[int(id)]
            network.architecture = arhitecture
            network.training_performance = training_performance
            network.validate_performance = validate_performance
            network.test_performance = test_performance
            network.training_error = training_erro
            network.validate_error = validate_error
            network.test_error = test_error
            network.learning_alghoritm = learning_alghoritm
            network.error_function = error_function
            network.activate1 = activate1
            network.activate2 = activate2
        else: 
            ConfigNotAllData = "Не удалось найти соответвие для некоторых сетей в теблице c конфигурацией"
            

    return networks

def get_dublicates(arg: tuple[any, any]) -> str: 
    dict_nets = dict(arg)
    out = ''
    for key in dict_nets: 
        count = 0
        for k, i in arg: 
            if k == key: 
                count += 1
        else: 
            if count > 1: 
                out += f"\n{key}"
    return out 

def get_networks_sesensitivity(ws: worksheet, networks) -> Mapping[id, Network]: 
    """Функция для чтения анализа чувствительности из файла xlsx"""
    parameters: tuple[str, ] | None = None
    row_len = 0 
    for row in list(map(list, ws.values)): 
        if parameters is None: 
            parameters = normal_row(row[1:])
            row_len = len(parameters)
            continue

        if (row[0] is not None) and (len(row[0].split('.')) == 2): 
            id, architecture, _ = *row[0].split('.'), ()
            sensitivity = tuple(zip_longest(parameters, row[1:row_len+1], fillvalue=None))
            
            if len(sensitivity) != len(dict(sensitivity)):
                print(len(sensitivity), len(dict(sensitivity)))
                raise IncorrectFile(f"В файле содержатся дубликаты: {get_dublicates(sensitivity)}")

            networks[int(id)] = Network(id, architecture, sensitivity)

        else: 
            continue
            
    return networks 


def normal_row(row: Sequence[any, ]) -> Sequence[any, ]: 
    out = [] 
    for item in row: 
        if item is None: 
            break  
        out += [item]

    return out 

if __name__ == "__main__": 
    networks = read('outt.xlsx')
    print(len(networks))